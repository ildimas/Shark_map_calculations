from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
import folium
from folium.features import DivIcon
from folium.plugins import MarkerCluster
import geocoder
import openpyxl
import geopy.distance
from mapmake.estate_price import PriceCalculation
from codes.forms import CodeForm
from mapmake.models import main_db, uploaded_files
from users.models import CustomUser
from dropzone.models import Doc

def index(request):
    # #Делаем карту
    m = folium.Map(location=[55.75, 37.62], height='100%', width='100%')
    folium.TileLayer('cartodbdark_matter', name='Темная').add_to(m)
    folium.TileLayer('cartodbpositron', name='Светлая').add_to(m)
    folium.TileLayer('stamentoner', name='Контрастная').add_to(m)
    # Кластеры
    cluster_new = MarkerCluster(name="Новые квартиры").add_to(m)
    cluster_old = MarkerCluster(name="Старые квартиры").add_to(m)
    cluster_uploaded = MarkerCluster(name='Загружено').add_to(m)
    heatmap_data = []
    #Разметка москвы
    layer_main_geo = folium.GeoJson(data=(open('./mapmake/geoJson/moscow.geojson', 'r', encoding='utf-8').read()), name='Районы москвы', show=False)
    folium.GeoJsonTooltip(fields=['name', 'cartodb_id']).add_to(layer_main_geo)
    layer_main_geo.add_to(m)
    #Данные на карту
    entire_db = main_db.objects.all()
    for i in range(len(entire_db)):
        some_db = entire_db[i] 
        lat = some_db.coordinates_lat
        lng = some_db.coordinates_lng
        if  lat > 56 or lng > 38 or lat < 55 or lng < 37:
            continue
        if some_db.estate_type == "вторичная":
            color_html = 'orange'
            marker = folium.Marker([lat, lng], icon=DivIcon(icon_size=(150,36), icon_anchor=(7,20), html=f'<span style="font-size: 25pt; color : {color_html}; style=display: block"> &#9660; </span>'), 
                        tooltip=f'{some_db.estate_type}', popup=f'{some_db.price, some_db.apartment_type, some_db.address ,some_db.estate_type, some_db.estate_type, f"Этаж квартиры: {some_db.flat_floor}",f"Количество комнат: {some_db.rooms_count}", f"Площадь квартиры: {some_db.main_square}"}',
                        )
            heatmap_data.append([some_db.coordinates_lat, some_db.coordinates_lng, some_db.price])
            
            
            cluster_old.add_child(marker)
        else:
            color_html = 'teal'
            marker = folium.Marker([lat, lng], icon=DivIcon(icon_size=(150,36), icon_anchor=(7,20), html=f'<span style="font-size: 25pt; color : {color_html}; style=display: block"> &#9660; </span>'), 
                        tooltip=f'{some_db.estate_type}', popup=f'{some_db.price, some_db.apartment_type, some_db.address ,some_db.estate_type, some_db.estate_type, f"Этаж квартиры: {some_db.flat_floor}",f"Количество комнат: {some_db.rooms_count}", f"Площадь квартиры: {some_db.main_square}"}',
                        )
            heatmap_data.append([some_db.coordinates_lat, some_db.coordinates_lng, some_db.price])
            
            cluster_new.add_child(marker)
        
    final_appartment_price_list = []
    entire_uploaded_db = uploaded_files.objects.all()
    main_bd = main_db.objects.all()
    for i in range(len(entire_uploaded_db)):
        fitable_obj_list = []; distances = []
        some_uploaded_db = entire_uploaded_db[i]
        target = [some_uploaded_db.coordinates_lat, some_uploaded_db.coordinates_lng]
        if (target[0] < 55) or (target[0] > 57) or (target[1] < 36) or (target[1] > 38):
            continue
        dict_of_appartments = {}
        for g in range(len(main_bd)):
            some_main_bd = main_bd[g]
            coords = [some_main_bd.coordinates_lat, some_main_bd.coordinates_lng]
            try:
                dist = geopy.distance.geodesic((coords[0], coords[1]), (target[0], target[1]))
                if dist <= 3:
                    fitable_obj_list.append(g)
                    distances.append(dist)
            except ValueError:
                continue
                
        for p in range(len(fitable_obj_list)):
            obj = main_bd[fitable_obj_list[p]]

            if obj.decor == 'нет ремонта':
                obj.decor = 'без отделки'
            if obj.apartment_type == 'not stated':
                obj.apartment_type = 'панельный'
            if obj.apartment_type.find('-') != -1:
                obj.apartment_type = obj.apartment_type[0: obj.apartment_type.find('-')]

            dict_of_appartments[f"{p}"] = [obj.estate_type, obj.price, obj.subway_distance, obj.rooms_count, obj.main_square,
                                        obj.kithcen_square, obj.flat_floor, obj.building_floor, obj.balcony, obj.decor, obj.apartment_type, distances[p]]
        
        main_object = [some_uploaded_db.estate_type, some_uploaded_db.subway_distance, some_uploaded_db.rooms_count,
        some_uploaded_db.main_square, some_uploaded_db.kithcen_square, some_uploaded_db.flat_floor, 
        some_uploaded_db.building_floor, some_uploaded_db.balcony, some_uploaded_db.decor, 
        some_uploaded_db.apartment_type]
        
        #квартира над которой проходит расчет
        # appartment_info[1] - расстояние до метро
        # appartment_info[2] - количество комнат
        # appartment_info[3] - площадь квартиры
        # appartment_info[4] - площадь кухни
        # appartment_info[5] - этаж квартиры
        # appartment_info[6] - этажность здания
    

        price_less_than_km = 0              # сумма цен квартир в пределах 1 км
        price_less_than_km_num = 0          # кол-во квартир в пределах 1 км
        price_more_than_km = 0              # сумма цен квартир в пределах 3 км
        price_more_than_km_num = 0          # кол-во квартир в пределах 3 км
        final_appartment_price = 0          # ФИНАЛЬНАЯ цена за квартиру   

        for key, value in dict_of_appartments.items():
            a = PriceCalculation(main_object, dict_of_appartments[key])
            if value[11] <= 1:
                price_less_than_km += value[1]
                price_less_than_km_num += 1
            else:
                price_more_than_km += value[1]
                price_more_than_km_num += 1
        try:
            if price_less_than_km_num == 0:
                final_appartment_price += price_more_than_km / price_more_than_km_num
            elif price_more_than_km_num == 0:
                final_appartment_price += price_less_than_km / price_less_than_km_num
            else:
                final_appartment_price += ((price_less_than_km / price_less_than_km_num) * 1.2 + (price_more_than_km / price_more_than_km_num) * 0.8) / 2
        except ZeroDivisionError:
            pass
        final_appartment_price = round(final_appartment_price * 0.955)
        print(final_appartment_price)
        final_appartment_price_list.append(final_appartment_price)
        
        upl_marker = folium.Marker([some_uploaded_db.coordinates_lat, some_uploaded_db.coordinates_lng], icon=DivIcon(icon_size=(150,36), icon_anchor=(7,20), html=f'<span style="font-size: 30pt; color : red ; style=display: block"> &#9660; </span>'), 
                    tooltip=f'{some_uploaded_db.estate_type}', popup=f'{some_uploaded_db.apartment_type, some_uploaded_db.address ,some_uploaded_db.estate_type, some_uploaded_db.estate_type, f"Этаж квартиры: {some_uploaded_db.flat_floor}",f"Количество комнат: {some_uploaded_db.rooms_count}", f"Площадь квартиры: {some_uploaded_db.main_square}", f"ЦЕНА:{final_appartment_price}"}')
        cluster_uploaded.add_child(upl_marker)
    entire_uploaded_db.delete()

    minimap = folium.plugins.MiniMap(toggle_display=True)
    m.add_child(minimap)
    folium.plugins.Fullscreen(position='topright').add_to(m)
    heatmap = folium.plugins.HeatMap(heatmap_data, radius=30, show=False, name='Тепловая карта стоимости жилья')
    heatmap.add_to(m)
    
    folium.LayerControl().add_to(m)
    m = m._repr_html_()
    context = {'m': m, 'fin_price': final_appartment_price_list,}
    return render(request, 'index.html', context=context)

def upload_view(request):
    if request.method == 'POST':
        try:
            my_file = request.FILES.get('file')
            Doc.objects.create(upload=my_file)
            print(request.POST)
            print(request.FILES)
            wb = openpyxl.load_workbook(my_file)
            ws = wb.active

            list_info = []
            for i in range(2, ws.max_row + 1):
                list_info.append([
                ws['A' + str(i)].value, ws['B' + str(i)].value, ws['C' + str(i)].value, ws['D' + str(i)].value,
                ws['D' + str(i)].value, ws['E' + str(i)].value, ws['F' + str(i)].value, ws['G' + str(i)].value, 
                ws['H' + str(i)].value, ws['I' + str(i)].value, ws['J' + str(i)].value, ws['K' + str(i)].value])
                
            u_db = uploaded_files.objects
            for i in range(len(list_info)):
                new_adress = 'Москва' + " " + (list_info[i])[0]
                y = geocoder.osm(new_adress)
                print(y.latlng)
                if y.latlng == None:
                    continue
                if (list_info[i])[1] == 'Студия':
                    rooms = 1
                else:
                    rooms = (list_info[i])[1]
                if (list_info[i])[8] == 'Да':
                    (list_info[i])[8] = True
                else:
                    (list_info[i])[8] = False
                u_db.create(address=new_adress, rooms_count=rooms, estate_type=(list_info[i])[2], building_floor=(list_info[i])[3]
                            ,apartment_type=(list_info[i])[5], flat_floor=(list_info[i])[6], main_square=(list_info[i])[7],
                            kithcen_square=15, balcony=(list_info[i])[8], subway_distance=(list_info[i])[10], decor=(list_info[i])[11],
                            coordinates_lng=(y.latlng)[1], coordinates_lat=(y.latlng)[0])
            return HttpResponse('upload')
        except BaseException:
            return JsonResponse('false')

def auth_view(request):
    form = AuthenticationForm()
    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            request.session['pk'] = user.pk
            return redirect('verify')
    return render(request, 'auth.html', {'form': form})

def verify_view(request):
    form = CodeForm(request.POST or None)
    pk = request.session.get('pk')
    if pk:
        user = CustomUser.objects.get(pk=pk)
        code = user.code
        code_user = f"{user.username}: {user.code}"
        if not request.POST:
            print(code_user)
            #sms sender
        if form.is_valid():
            num = form.cleaned_data.get('number')
            if str(code) == num:
                code.save()
                login(request, user)
                return redirect('home')
            else:
                return redirect('login')
    return render(request, 'verify.html', {'form': form})